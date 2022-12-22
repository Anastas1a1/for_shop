from asyncio import FastChildWatcher
import pandas as pd
from pandas.io.excel import ExcelWriter
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import namedtuple
import os
import os.path



def get_names(us_id):
    path = 'database/' + str(us_id) + '.xlsx' 
    if os.path.exists(path) == False:
        return 0
    names_list = []
    wb = load_workbook(path)
    sheet = wb.worksheets[0] 
    row=2       
    while sheet.cell(row = row, column=1).value!=None:
        names_list.append(sheet.cell(row = row, column=1).value)
        row+=1
    if row == 2:
        return 0
    return names_list



def add_transaction_1(us_id, transaction):
    path = 'database/' + str(us_id) + '.xlsx' 
    if os.path.exists(path) == False:
        wb = Workbook()
        ws=wb.active
        col =['name','fee total:', '0.00', 'note']
        ws.append(col)
        wb.save(path)
        
    wb = load_workbook(path)
    sheet = wb.worksheets[0]
 
    column = 1
    col_fee = 1
    col_rate = 1
    row = 1
    while sheet.cell(row = row, column=column).value!=None :
        if  sheet.cell(row = row, column=column).value==transaction['name']:
            break
        print(sheet.cell(row = row, column=column).value)
        row+=1
    sheet.cell(row = row, column=column).value = transaction['name']   
     
    
    
    while sheet.cell(row = 1, column=column).value!='fee total:':
        if sheet.cell(row = 1, column=column).value == transaction['cur']:
            col_rate=column
        column+=1       
    col_fee=column
    
    column+=2 
    sheet.cell(row = row, column=column).value = transaction['note']  
      
    if col_rate == 1:
        sheet.insert_cols(col_fee, 1)
        col_rate=col_fee

    
    sheet.cell(row=1, column=col_rate).value = transaction['cur']
    if sheet.cell(row=row, column=col_rate).value != None:
        sheet.cell(row=row, column=col_rate).value += transaction['sum']
    else:
        sheet.cell(row=row, column=col_rate).value = transaction['sum']

    wb.save(path)
    wb.close() 


# add_transaction_1(1065409297, {'name': 'Tor', 'rate': 'EUR', 'sum': 13465, 'note': 'Hi'})
